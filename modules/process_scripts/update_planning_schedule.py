# file_name:   _update_planning_schedule.py
# created_on:  2023-08-17 ; vicente.diaz ; updated self.state_name attribute
# modified_on: 2023-09-09 ; santiago.garcia

import os
import sys
import logging
sys.path.append(os.path.abspath(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))  # move to modules 
from _fmw.fmw_utils import *
from _fmw.fmw_classes import BusinessException
from classes.robot_date import RobotDate
from process_scripts._base_process_class import ProcessBase
from openpyxl import load_workbook


class UpdatePlanningSchedule(ProcessBase):
    def __init__(self, config:dict):
        ProcessBase.__init__(self, config=config) 
        self.state_name = self.state_name = type(self).__name__ # Get the class name, do not change     
        # workflow parameters 
        self.process_input_file = self.config_global["PROCESS_INPUT_FILE"]
        self.process_input_file_worksheet = self.config_global["PROCESS_INPUT_FILE_WORKSHEET"]
        self.worktray_template = self.config_global["WORKTRAY_TEMPLATE"]
        self.worktray_file = self.config_global["WORKTRAY_FILE"]
        self.process_data_files_worksheet = self.config_global["PROCESS_DATA_FILES_WORKSHEET"]
        self.process_input_file_name = os.path.basename(self.config_global["PROCESS_INPUT_FILE"])
        self.clickup_data_path = self.config_global["CLICKUP_DATA_PATH"]
        self.processed_file_path = self.config_global["PROCESSED_FILE_PATH"]
        self.raw_files_path = self.config_global["RAW_FILE_PATH"]
        self.worktray_filter_column = self.config_global["WORKTRAY_FILTER_COLUMN"]
        self.process_data_file_timestamp = self.config_global["PROCESS_DATA_FILE_TIMESTAMP_FORMAT"]
        # These variables already exist by inheritance
        # self.config              = config
        # self.environment         = config["METADATA"]["ENVIRONMENT"].upper()
        # self.config_env          = config[self.environment]
        # self.config_global       = config["GLOBAL"]
        # self.config_fmw          = config["FRAMEWORK"]        
        # self.process_data        = self.config_fmw["PROCESS_DATA"]
        # self.config_emails       = self.config["EMAIL"]
        # self.now                 = dt.datetime.now()

    def generate_worktray(self):
        logging.info(f"--- GENERATING THE WORKTRAY ---")
        PROCESS_INPUT_FILE = self.process_input_file
        PROCESS_INPUT_FILE_WORKSHEET = self.process_input_file_worksheet
        WORKTRAY_TEMPLATE = self.worktray_template
        WORKTRAY_FILE = self.worktray_file
        PROCESS_DATA_FILES_WORKSHEET = self.process_data_files_worksheet
        PROCESS_INPUT_FILE_NAME = self.process_input_file_name
        CLICKUP_DATA_PATH = self.clickup_data_path
        PROCESSED_FILE_PATH = self.processed_file_path
        PROCESS_DATA_FOLDER = self.process_data
        RAW_FILES_PATH = self.raw_files_path
        WORKTRAY_FILTER_COLUMN = self.worktray_filter_column
        logging.info(f"Input file name: {PROCESS_INPUT_FILE}")
        #The input file structure must be the same as the worktray (it's memory file actually, but we will call it the "master" worktray)
        shutil.copy(WORKTRAY_TEMPLATE, WORKTRAY_FILE) #We generate the pretty template structure of the file (because the writer add data inside a new sheet independant of the format)
        #Also, we want to make shure that this file exist. Otherwise, the process can't continue.
        if not os.path.exists(PROCESS_INPUT_FILE):
             raise BusinessException(f"The input file '{PROCESS_INPUT_FILE_NAME}' isn't located in the correct directory!")
        #We will read the excel input file and convert it into a dataframe
        df = pd.read_excel(PROCESS_INPUT_FILE, sheet_name= PROCESS_INPUT_FILE_WORKSHEET)
        #We keep only the rows that we need to process
        worktray = df[df[WORKTRAY_FILTER_COLUMN] == 'FALSE']
        worktray_columns = tuple(worktray.columns) #Tuple because this structure will never changes in order and data.
        total_rows = len(worktray)
        for row_idx in range(total_rows):
            execution_datetime = now()
            timestamp = execution_datetime.to_string(process_data_file_timestamp)
            #CHECKPOINT I NEED TO ADD ALL THE REMAINING PATHS OF THE PROCESS
            project_id = worktray.iloc[row_idx, 0]
            project_finished = worktray.iloc[row_idx, 1]
            planning_data_extracted = worktray.iloc[row_idx, 2]
            clickup_data_extracted = worktray.iloc[row_idx, 3]
            clickup_data_path = "{CLICKUP_DATA_PATH}".format(project_id, timestamp)
            raw_file_path = "{RAW_FILES_PATH}".format(project_id, timestamp) 
            processed_file_path = "{PROCESSED_FILE_PATH}".format(project_id, timestamp)
            planning_file_updated = "{PLANNING_FILE_UPDATED}".format(project_id, timestamp)
            # execution_datetime (already created)
            observations = None # We will add this later in the next process steps
            #Now we insert this data into the dataframe
            insert_data = []
            insert_data.append(project_id)
            insert_data.append(project_finished)
            insert_data.append(planning_data_extracted)
            insert_data.append(clickup_data_extracted)
            insert_data.append(clickup_data_path)
            insert_data.append(raw_file_path)
            insert_data.append(processed_file_path)
            insert_data.append(planning_file_updated)
            insert_data.append(execution_datetime)
            insert_data.append(observations)
        #After the insert data is ready, we will insert it into the worktray
        book = load_workbook(PROCESS_INPUT_FILE)
        #We need to ensure that the worksheet exists, otherwise we will create it
        sheet = book[PROCESS_INPUT_FILE_WORKSHEET] if PROCESS_INPUT_FILE_WORKSHEET in book.sheetnames else book.create_sheet(PROCESS_INPUT_FILE_WORKSHEET)
        data = worktray.values.tolist()
        for row_data in data:
             sheet.append(row_data)
        book.save(WORKTRAY_FILE)
        logging.info(f"Reached example step 1 {PROCESS_INPUT_FILE}")
        logging.info("WORKTRAY_FILE GENERATED SUCCESFULLY")
        return 0
    
    def extract_clickup_project_data(self):
        logging.info(f"--- DESCRIPTIVE START FOR SPRIPT 2 ---")
        # Run script after this
        # raise BusinessException("Testing business exceptions") # Testing business exception
        logging.info(f"Reached example step 2 {self.template_parameter_2}")
        logging.info(f"DESCRIPTIVE END FOR SPRIPT 2")
        return 0

    def run_workflow(self):
        logging.info(f"----- Starting state: {self.state_name} -----")
        try: # Add workflow in try block bellow
            self.generate_worktray()        
            #self.script_function_2()
        except BusinessException as error:
            self._build_business_exception(error)
            raise error
        except Exception as error:
            raise error  
        logging.info(f"Finished state: {self.state_name}")

if __name__ == "__main__":
    start_logging(logs_level="DEBUG", show_console_logs=True, save_logs=False)
    config = read_config()
    logging.info(f"--------- Script {os.path.basename(__file__)} started ---------")    
    state = UpdatePlanningSchedule(config=config)
    state.run_workflow()
    # state.script_function_1()
    logging.info(f"--------- Script {os.path.basename(__file__)} finished ---------")     
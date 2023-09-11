# file_name:   fmw_utils.py
# created_on:  2023-01-16 ; vicente.diaz
# modified_on: 2023-03-14 ; vicente.diaz ; guillermo.konig
# modified_on: 2023-04-18 ; guillermo.konig
# modified_on: 2023-06-08 ; vicente.diaz ; fix read config with env tags
# modified_on: 2023-06-15 ; vicente.diaz ; added remove_path_until_tag() function
# modified_on: 2023-07-18 ; guillermo.konig ; updated save_excel_file to include overwrite option
# modified_on: 2023-07-18 ; guillermo.konig ; added read and write text and remove tildes
# modified_on: 2023-08-17 ; vicente.diaz ; added function number_to_currency()

import re
import io
import os
import json
import shutil
import logging
import subprocess
import unicodedata
import pandas as pd
import datetime as dt
from dateutil.relativedelta import relativedelta
import openpyxl
from openpyxl.styles import Font, PatternFill

def start_logging(log_folder="", logs_level="INFO", show_console_logs=True, save_logs=False, 
                  label="", timestamp_format="%Y%m%d_%H%M%S"):
    logFormatter_file    = logging.Formatter('%(asctime)-15s [%(levelname)-5.5s] %(message)s')
    logFormatter_console = logFormatter_file
    now_str              = dt.datetime.now()
    logs_now_filepath    = os.path.join(log_folder,f"{label}_{now_str.strftime(timestamp_format)}.log")
    # assign root logger
    rootLogger = logging.getLogger()
    rootLogger.setLevel(logs_level)
    if show_console_logs:
        # show logs on console
        consoleHandler = logging.StreamHandler()
        consoleHandler.setFormatter(logFormatter_console)
        rootLogger.addHandler(consoleHandler)
    if save_logs:
        create_folder(folder_path=log_folder)
        # write logs in log file
        fileHandler = logging.FileHandler(logs_now_filepath, mode="a")
        fileHandler.setFormatter(logFormatter_file)
        rootLogger.addHandler(fileHandler)
    logging.info(f"--------- Logging started ---------")
    if save_logs:
        logging.info(f"Log file: {logs_now_filepath}")
    return 0


class Config:
    def __init__(self):
        self.execution_folder = os.path.dirname(os.path.dirname(os.path.dirname(__file__)))
        self.config   = None
        self.tag_keys = dict()

    def build_config(self):
        # utils.py must be in modules folder
        # Replace auto user in all paths
        self.config = self.read_config()
        self.config = {k.upper(): v for k, v in self.config.items()}
        self.config = self.build_absolute_config_paths(config=self.config) # Create absolute paths
        self.get_config_tag_keys(config=self.config)                       # Create tags keys dict
        config_env  = self.config[self.config["METADATA"]["ENVIRONMENT"]]  # Assig env dict
        self.get_config_tag_keys(config=config_env)                        # Update tags keys dict with env dict
        self.config = self.build_absolute_config_paths(config=self.config) # Update values with env tag keys
        self.config = self.build_absolute_config_paths(config=self.config) # Update values with env tag keys
        self.config = self.build_absolute_config_paths(config=self.config) # Update values with env tag keys
        return self.config
        
    def read_config(self):
        config_file = os.path.join(self.execution_folder, "config.jsonc")
        if os.path.exists(config_file):
            config = read_jsonc(file_path=config_file)
            config = {k.upper(): v for k, v in config.items()}
            return config
        else: 
            raise Exception("Could not read config file: utils.py must be located in modules/fmw folder and config.json in framework folder")

    def get_config_tag_keys(self, config:dict):
        for key, value in config.items():
            if type(value) is dict:
                self.get_config_tag_keys(value)
            elif type(value) is str:
                self.tag_keys[key] = value
        
    def build_absolute_config_paths(self, config:dict):
        for key, value in config.items():
            if type(value) is dict:
                value = self.build_absolute_config_paths(value)
            elif type(value) is str:
                if value.startswith("../"):
                    value_no_code = os.path.normpath(value[3:]) # remove "../"
                    value = os.path.join(self.execution_folder,value_no_code)
                if "\\Users\\" in value:
                    path_vals = value.split("\\")
                    user_idx  = path_vals.index("Users") + 1
                    value     = value.replace(path_vals[user_idx], os.getlogin())
                elif "/Users/" in value:
                    path_vals = value.split("/")
                    user_idx  = path_vals.index("Users") + 1
                    value     = value.replace(path_vals[user_idx], os.getlogin())
                tags_in_value = list(filter(lambda tag_key: "{"+tag_key+"}" in value, self.tag_keys.keys()))
                # print(f"{key}: {value} | {tags_in_value}")
                if len(tags_in_value) > 0:
                    value = value.replace("{"+tags_in_value[0]+"}", self.tag_keys[tags_in_value[0]])
                value = value.replace("{WORKING_FOLDER}", f"{os.getcwd()}/")
                config[key] = value.replace("{USER_PROFILE}", os.environ['USERPROFILE'])
        return config
    

def read_config():
    config_class = Config()
    config = config_class.build_config()
    return config
    

def read_json(file_path: str):
    with open(file_path, "r", encoding="utf-8") as f:
        return json.load(f)
    

def read_jsonc(file_path: str):
    with open(file_path, "r", encoding="utf-8") as f:
        file_content = f.read()
    file_content = re.sub(r'\/\/.*?\n|\/\*.*?\*\/', '', file_content, flags=re.DOTALL)
    file_obj = io.StringIO(file_content)
    return json.load(file_obj)


def save_json_file(dicData: dict, file_path: str):
    print(f"Saving json file as: {file_path}")
    with open(file_path, "w") as file:
        json.dump(dicData, file, indent=4, sort_keys=True)


def delete_folder(folder_path: str):
    logging.info(f"Deleting folder: {folder_path}")
    if os.path.exists(folder_path):
        shutil.rmtree(folder_path, ignore_errors=False, onerror=None)


def create_folder(folder_path: str):
    logging.info(f"Creating folder {folder_path}")
    if not os.path.exists(folder_path):
        # Create the folder
        os.makedirs(folder_path, exist_ok=True)


def save_excel_file(df:pd.DataFrame, file_path:str, sheet_name:str="base", overwrite:bool=False, index:bool=False):
    logging.info(f"Saving dataframe with shape: {df.shape}")
    # Create the file if not already exists
    if not os.path.exists(file_path) or overwrite:
        writer = pd.ExcelWriter(file_path, engine='xlsxwriter')
        df.to_excel(writer, sheet_name=sheet_name, index=index)  # send df to writer
        worksheet = writer.sheets[sheet_name]                    # pull worksheet object
        for idx, col in enumerate(df):                           # loop through all columns
            series  = df[col]
            max_len = min(max(series.astype(str).map(len).max(), len(str(series.name))) + 3, 50)
            if pd.isna(max_len):
                max_len = len(col) + 3
            worksheet.set_column(idx, idx, max_len)             # set column
        writer.close()
    else:
        # save the dataframe to the same excel file
        with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists="overlay") as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=index)
        logging.info(f"File saved as {file_path}")
    return 0

def reset_worktray_headers_format(worktray_path: str, sheet_name: str, worktray_template: str):
    # Open the worktray and template workbooks
    workbook = openpyxl.load_workbook(worktray_path)
    worksheet = workbook[sheet_name]
    template_workbook = openpyxl.load_workbook(worktray_template)
    template_worksheet = template_workbook[sheet_name]
    # Loop through each cell in the worksheet and apply template formatting
    for col_idx in range(1, worksheet.max_column + 1):
        for row_idx in range(1, worksheet.max_row + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            template_cell = template_worksheet.cell(row=row_idx, column=col_idx)
            # Create new Font and PatternFill objects based on the template
            new_font = Font(
                name=template_cell.font.name,
                bold=template_cell.font.bold,
                italic=template_cell.font.italic,
                color=template_cell.font.color,
            )
            new_fill = PatternFill(
                start_color=template_cell.fill.start_color,
                end_color=template_cell.fill.end_color,
                fill_type=template_cell.fill.fill_type,
            )
            # Apply the new Font and PatternFill to the cell
            cell.font = new_font
            cell.fill = new_fill
    # Save the modified workbook with the same name as the original
    workbook.save(worktray_path)

    return 0

def kill_processes(processes:list):
    processes = [process.lower().strip() for process in processes]
    logging.info(f"Killing process: {processes}")
    for proc in processes:
        logging.debug(f"Killing process: {proc}")
        process_status = subprocess.run(["powershell", "-command", f"taskkill /IM {proc} /F"]
                                        , capture_output=True)
        logging.debug(f"Process {proc} killed")
    return 0     


def add_dt_offset(in_dt:dt.datetime=dt.datetime.now(), months:int=0, years:int=0, days:int=0, hours:int=0, minutes:int=0, seconds:int=0):
    return in_dt + relativedelta(years=years, months=months, days=days, hours=hours, minutes=minutes, seconds=seconds)


def last_day_in_month(in_dt:dt.datetime):
    next_month = add_dt_offset(in_dt=in_dt, month=1).replace(day=1)
    last_day   = add_dt_offset(in_dt=next_month, days=-1)
    return last_day


def dt_to_spanish(in_dt: dt.datetime, month_or_day:str="month", capitalize:bool=False):
    months_dict = {
        1: 'enero',
        2: 'febrero',
        3: 'marzo',
        4: 'abril',
        5: 'mayo',
        6: 'junio',
        7: 'julio',
        8: 'agosto',
        9: 'septiembre',
        10: 'octubre',
        11: 'noviembre',
        12: 'diciembre'
    }
    days_dict = {
        0: 'lunes',
        1: 'martes',
        2: 'miercoles',
        3: 'jueves',
        4: 'viernes',
        5: 'sabado',
        6: 'domingo'
    }
    if month_or_day.lower() == "month":
        month_num = int(in_dt.month)
        output    = months_dict[month_num]
    elif month_or_day.lower() == "day":
        day_num   = in_dt.weekday()
        output    = days_dict[day_num]
    else:
        raise Exception(f"Input to month_or_day variable: {month_or_day} does not match expected answers")
    if capitalize:
        output = output.capitalize()
    return output


def log_worktray_metadata(config: dict, wt_status_col:str="STATUS"):
    logging.info("--- Logging worktray metadata ---")
    global_keys = config["GLOBAL"].keys()
    process_with_worktray = any(["WORKTRAY" in key.upper() for key in global_keys])
    logging.info(f"Process with worktray: {process_with_worktray}")
    process_data_path = config["FRAMEWORK"]["PROCESS_DATA"]
    worktray_path = os.path.join(process_data_path, "worktray.xlsx")
    logging.info(f"Worktray path: {worktray_path}")
    worktray_exists = os.path.exists(worktray_path)
    logging.info(f"Worktray exists: {worktray_exists}")
    wt_status = dict()
    wt_rows   = 0
    if process_with_worktray:
        if worktray_exists:
            df_worktray = pd.read_excel(io=worktray_path)
            logging.info(f"Worktray shape: {df_worktray.shape}")
            wt_rows   = df_worktray.shape[0]
            wt_status = df_worktray[wt_status_col].value_counts().to_dict()
    else:
        wt_rows = 1
    logging.info(f"Worktray rows: {wt_rows}")
    logging.info(f"Worktray status: {wt_status}")
    return 0


def get_total_states():
    execution_folder = os.path.dirname(os.path.dirname(os.path.dirname(__file__)))
    main_file        = os.path.join(execution_folder, "main.py")
    with open(main_file, "r", encoding="utf-8") as f:
        main_file_text = f.read().splitlines()
    state_lines = [line for line in main_file_text if "self.state_idx == " in line]
    return len(state_lines)


def mask_number(nfloat, digits_to_show, decimals_to_show):
    #nfloat = round(nfloat, ndigits=decimals_to_show)    
    number = "{:.{}f}".format(nfloat, decimals_to_show )
    total_digits_to_show = digits_to_show + (0 if decimals_to_show == 0 else decimals_to_show + 1)
    mask = f"{'*' * (len(number) - total_digits_to_show)}{number[-(total_digits_to_show):]}"
    return (mask)


def get_size(file_path, unit='bytes'):
        file_size = os.path.getsize(file_path)
        exponents_map = { 'bytes': 0, 'kb': 1, 'mb': 2, 'gb': 3 }        
        if unit not in exponents_map:
            raise ValueError("Must select from \
            ['bytes', 'kb', 'mb', 'gb']")
        else:
            size = file_size / 1024 ** exponents_map[unit]
            return round(size, 3)


def get_last_log_file(log_folder:str):
    log_files = [os.path.join(log_folder, file_name) for file_name in os.listdir(path=log_folder)]
    log_files.sort(reverse=True)
    last_log_file = log_files[0]
    logging.info(f"Execution log file: {os.path.basename(last_log_file)}")
    return last_log_file


def build_execution_summary_table(config:dict, file_exe_report, dti_start_exe=dt.datetime.now(), dti_final_exe=dt.datetime.now(), 
                                  status=None, processed_states=None, total_state=None):
    dic_exe_data = dict()
    # Add execution metadata
    dic_exe_data["EXECUTION_STATUS"] = status
    dic_exe_data["PROCESSED_STATES"] = processed_states
    dic_exe_data["TOTAL_STATES"]     = total_state
    dic_exe_data["START_TIME"]       = dti_start_exe.strftime("%Y-%m-%d_%H:%M:%S")
    dic_exe_data["END_TIME"]         = dti_final_exe.strftime("%Y-%m-%d_%H:%M:%S")
    # Add process METADATA
    dic_exe_data.update(config["METADATA"])
    # Convert dictionary to dataframe
    df_exe_report = pd.DataFrame(data=dic_exe_data.items(), columns=["PARAMETER", "VALUE"])
    # Save dataframe as json file
    save_json_file(dicData=dic_exe_data, file_path=file_exe_report)
    return df_exe_report


def remove_path_until_tag(path:str, tag:str, include_tag:bool=True) -> str:
    """
    Function to shorten a given path, removing its first part until a specific folder.
     Can be used to remove the local part of a path, improving readability for the user.  
    - path: path to be shortened. 
    - tag: target folder (name or path) to split the path. 
    - include_tag: if True, the path starts with the tag folder; if False, the path starts with the subfolder of the tag
    """
    path_parts  = []
    found_tag   = False
    out_path    = path
    path_exists = True
    while path_exists:
        head, tail = os.path.split(path)
        path_parts.insert(0, tail)
        if tag in [head, tail]:
            found_tag = True
            if head == tag:
                path_parts.insert(0, os.path.split(head)[-1])
            break
        path = head
        path_exists = head and tail
    if found_tag:
        if not include_tag:
            path_parts.pop(0)  # Remove the tag from the path
        out_path = os.path.join(*path_parts)
    return out_path

def save_txt_file(input_txt:str, out_path:str):
    create_folder(os.path.dirname(out_path))
    f = open(out_path, "w", encoding="utf-8")
    f.write(input_txt)
    f.close()
    return 0

def read_txt(file_path: str):
    with open(file_path, "r", encoding="utf-8") as f:
        return f.read()

def remove_accents(string:str) -> str:
    """ Remove accents of string"""
    nfkd_form = unicodedata.normalize('NFKD', string)
    not_accents_string = ''.join([c for c in nfkd_form if not unicodedata.combining(c)])
    return not_accents_string


def number_to_currency(number:float, type_format:str="currency", sign:bool=True) -> str:
        """ Convert number to chilean currency
            Parameters:
                number: Number to convert to chilean currency
                type_format: Currency format type to convert [dots_only, two_decimal, currency]
                sign: Add the weight sign ($) to the formatted number [True, False]
            Example:
                dots_only:   10000 to 10.000    | If sign = True: 10000 to $10.000
                two_decimal: 10000 to 10000,00  | If sign = True: 10000 to $10000,00
                currency:    10000 to 10.000,00 | If sign = True: 10000 to $10.000,00
        """
        if type_format == "dots_only":
            currency_string = "{:,}".format(number)
        elif type_format == "two_decimal":
            currency_string = "{:.2f}".format(number)
        elif type_format == "currency":
            currency_string = "{:,.2f}".format(number)
        if sign:
            currency_string = "$" + currency_string          
        currency_string_spanish = currency_string.replace(".", ";"). replace(",", ".").replace(";", ",")
        return currency_string_spanish


if __name__ == "__main__":
    start_logging()
    config = read_config()
    print("Config")
    for k, v in config.items():
        print(f"{k}: {v}")
        print()


